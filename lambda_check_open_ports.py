import boto3
import openpyxl
import logging
import os
import json
from botocore.exceptions import ClientError


logger = logging.getLogger()
logger.setLevel(logging.INFO)

s3_client = boto3.client('s3')
ec2_client = boto3.client('ec2')


def lambda_handler(event, context):
    try:
        # Retrieve all security groups
        response = ec2_client.describe_security_groups()
        security_groups = response['SecurityGroups']
        
        # Filter the security groups that have open port 22 (SSH)
        filtered_security_groups = [sg for sg in security_groups if is_port_open(sg, 22)]
        
        # Create a new Excel workbook
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        
        # Write headers to worksheet
        worksheet.cell(row=1, column=1, value="Group Name")
        worksheet.cell(row=1, column=2, value="Description")
        worksheet.cell(row=1, column=3, value="VPC ID")
        worksheet.cell(row=1, column=4, value="Inbound Rule")
        
        # Write security groups details to worksheet
        for i, sg in enumerate(filtered_security_groups):
            worksheet.cell(row=i+2, column=1, value=sg['GroupName'])
            worksheet.cell(row=i+2, column=2, value=sg['Description'])
            worksheet.cell(row=i+2, column=3, value=sg['VpcId'])
            worksheet.cell(row=i+2, column=4, value=json.dumps(sg['IpPermissions'], indent=4))
            
        # Save Excel workbook to S3 bucket
        file_name = os.environ['OPEN_PORTS_FILE_NAME']
        workbook.save('/tmp/'+file_name)
        bucket_name = os.environ['BUCKET_NAME']
        s3_client.upload_file('/tmp/'+file_name, bucket_name, file_name)
        os.remove('/tmp/'+file_name)
        
        return {
            'statusCode': 200,
            'body': 'Excel file successfully created and uploaded to S3 bucket'
        }
    except ClientError as e:
        logger.error(e)
        return {
            'statusCode': e.response['ResponseMetadata']['HTTPStatusCode'],
            'body': e.response['Error']['Message']
        }
    except Exception as e:
        logger.error(e)
        return {
            'statusCode': 500,
            'body': 'Internal server error'
        }


def is_port_open(security_group, port):
    """Returns True if the security group allows incoming traffic on the specified port"""
    for rule in security_group['IpPermissions']:
        if 'FromPort' in rule and rule['FromPort'] <= port and 'ToPort' in rule and rule['ToPort'] >= port:
            for ip_range in rule['IpRanges']:
                if ip_range['CidrIp'] == '0.0.0.0/0':
                    return True
    return False
